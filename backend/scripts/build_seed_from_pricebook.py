"""Generate app/seed.py from the Exicom pricebook Excel.

Run this whenever the pricebook changes:
    python scripts/build_seed_from_pricebook.py "path/to/Pricebook.xlsx"

It produces backend/app/seed.py with every product and a full
multi-currency / multi-tier (MoQ) price matrix, so the order form can
re-price lines when the currency or quantity changes.
"""
from __future__ import annotations

import sys
import re
from pathlib import Path

import openpyxl

HERE = Path(__file__).resolve().parent
SEED_PATH = HERE.parent / "app" / "seed.py"

# Currency preference order when choosing a product's default price.
CURRENCY_ORDER = ["USD", "EUR", "INR", "MYR"]

# Tier brackets per sheet (min, max) — max None = open ended.
AC_TIERS = [(1, 50), (51, 150), (151, None)]
DC_TIERS = [(1, 5), (6, 10), (11, None)]      # USD / EUR / MYR brackets
DC_TIERS_INR = [(1, 2), (3, 5), (6, None)]    # INR brackets differ on DC sheet


def clean(text) -> str:
    if text is None:
        return ""
    return str(text).replace("–", "-").replace("�", "-").strip()


def num(val):
    if val is None or (isinstance(val, str) and not val.strip()):
        return None
    try:
        return round(float(val), 2)
    except (TypeError, ValueError):
        return None


def slugify_code(name: str, category: str) -> str:
    """Build a synthetic part code for rows the pricebook leaves blank."""
    prefix = {"AC Charger": "AC", "DC Charger": "DC", "Accessories": "ACC"}.get(category, "PRD")
    kw = re.search(r"(\d+)\s*kw", name, re.IGNORECASE)
    amp = re.search(r"(\d+)\s*a(?:mps?)?\b", name, re.IGNORECASE)
    bits = [prefix]
    if kw:
        bits.append(f"{kw.group(1)}K")
    if amp:
        bits.append(f"{amp.group(1)}A")
    tail = "NDLS" if "NDLS" in name.upper() else ("DLS" if "DLS" in name.upper() else "")
    if tail:
        bits.append(tail)
    return "-".join(bits)


def build_tiers(values: list, brackets: list) -> list | None:
    """Pair price values with bracket ranges, dropping empty cells."""
    tiers = []
    for (lo, hi), v in zip(brackets, values):
        p = num(v)
        if p is not None:
            tiers.append([lo, hi, p])
    if not tiers:
        return None
    # Collapse to a single open tier if every bracket has the same price.
    if len({t[2] for t in tiers}) == 1:
        return [[1, None, tiers[0][2]]]
    return tiers


def first_tier_price(prices: dict) -> tuple[float, str]:
    for cur in CURRENCY_ORDER:
        if cur in prices and prices[cur]:
            return prices[cur][0][2], cur
    return 0.0, "USD"


def parse_ac(ws) -> list[dict]:
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        code, name = clean(row[0]), clean(row[1])
        if not name:
            continue
        prices = {}
        for cur, cols in (("USD", row[2:5]), ("INR", row[5:8]), ("EUR", row[8:11]), ("MYR", row[11:14])):
            t = build_tiers(list(cols), AC_TIERS)
            if t:
                prices[cur] = t
        out.append(make_product(code, name, prices, "AC Charger"))
    return out


def parse_dc(ws) -> list[dict]:
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        code, name = clean(row[0]), clean(row[1])
        if not name:
            continue
        prices = {}
        for cur, cols, brackets in (
            ("USD", row[2:5], DC_TIERS),
            ("INR", row[5:8], DC_TIERS_INR),
            ("EUR", row[8:11], DC_TIERS),
            ("MYR", row[11:14], DC_TIERS),
        ):
            t = build_tiers(list(cols), brackets)
            if t:
                prices[cur] = t
        out.append(make_product(code, name, prices, "DC Charger"))
    return out


def parse_accessories(ws) -> list[dict]:
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        code, name = clean(row[0]), clean(row[1])
        if not name:
            continue
        prices = {}
        for cur, val in (("USD", row[2]), ("INR", row[3]), ("EUR", row[4]), ("MYR", row[5])):
            p = num(val)
            if p is not None:
                prices[cur] = [[1, None, p]]
        out.append(make_product(code, name, prices, "Accessories"))
    return out


def make_product(code: str, name: str, prices: dict, category: str) -> dict:
    # Multi-line pricebook names carry spec lines after the first line.
    parts = [p.strip() for p in name.split("\n") if p.strip()]
    product_name = parts[0]
    description = "\n".join(parts[1:])
    if not code:
        code = slugify_code(product_name, category)
    unit_price, currency = first_tier_price(prices)
    return dict(
        product_code=code,
        code_note=category,
        product_name=product_name,
        description=description,
        unit_price=unit_price,
        currency=currency,
        unit="Nos.",
        category=category,
        prices=prices,
    )


def render_seed(products: list[dict]) -> str:
    lines = [
        '"""Seed the catalog from the Exicom pricebook.',
        "",
        "AUTO-GENERATED by scripts/build_seed_from_pricebook.py — do not edit by hand.",
        "Each product carries a multi-currency / multi-tier (MoQ) price matrix in",
        "`prices`: { currency: [[min_qty, max_qty_or_None, price], ...] }.",
        "",
        "Run idempotently:        python -m app.seed",
        "Wipe & reload catalog:   python -m app.seed --reset",
        '"""',
        "import sys",
        "",
        "from app.database import SessionLocal, Base, engine",
        "from app import models, migrate",
        "",
        "SEED_PRODUCTS = [",
    ]
    for p in products:
        lines.append("    dict(")
        lines.append(f"        product_code={p['product_code']!r},")
        lines.append(f"        code_note={p['code_note']!r},")
        lines.append(f"        product_name={p['product_name']!r},")
        lines.append(f"        description={p['description']!r},")
        lines.append(f"        unit_price={p['unit_price']!r},")
        lines.append(f"        currency={p['currency']!r},")
        lines.append(f"        unit={p['unit']!r},")
        lines.append(f"        category={p['category']!r},")
        lines.append(f"        prices={p['prices']!r},")
        lines.append("    ),")
    lines.append("]")
    lines.append("")
    lines.append(SEED_MAIN)
    return "\n".join(lines) + "\n"


SEED_MAIN = '''
def main(reset: bool = False) -> None:
    migrate.run()  # ensure new columns exist on already-deployed databases
    Base.metadata.create_all(bind=engine)
    db = SessionLocal()
    try:
        if reset:
            deleted = db.query(models.CatalogProduct).delete()
            db.commit()
            print(f"Reset: removed {deleted} existing catalog products.")

        inserted = updated = 0
        for p in SEED_PRODUCTS:
            obj = (
                db.query(models.CatalogProduct)
                .filter_by(product_code=p["product_code"], product_name=p["product_name"])
                .first()
            )
            if obj:
                for k, v in p.items():
                    setattr(obj, k, v)
                updated += 1
            else:
                db.add(models.CatalogProduct(**p))
                inserted += 1
        db.commit()
        print(f"Catalog seed complete: {inserted} inserted, {updated} updated.")
    finally:
        db.close()


if __name__ == "__main__":
    main(reset="--reset" in sys.argv)
'''


def main():
    xlsx = sys.argv[1] if len(sys.argv) > 1 else None
    if not xlsx:
        # default: the pricebook in the project root
        candidates = list((HERE.parent.parent.parent).glob("*Pricebook*.xlsx"))
        if not candidates:
            sys.exit("Pass the pricebook .xlsx path as the first argument.")
        xlsx = str(candidates[0])

    print(f"Reading {xlsx}")
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    products = []
    products += parse_ac(wb["AC Charger"])
    products += parse_dc(wb["DC Charger"])
    products += parse_accessories(wb["Accessories Spare"])

    SEED_PATH.write_text(render_seed(products), encoding="utf-8")
    print(f"Wrote {len(products)} products to {SEED_PATH}")


if __name__ == "__main__":
    main()
